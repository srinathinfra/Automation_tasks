"""
build_gap_report.py
-------------------
The main script. Run this.

  python build_gap_report.py --from 2026-07-01 --to 2026-08-31

What it does:
  1. Asks Jira what you have ALREADY logged, day by day.
  2. Reads your exported Outlook calendar for the same window.
  3. Works out, for each working day, how many hours are still missing.
  4. Writes an Excel workbook whose last sheet is ready to feed into the
     script you already have for pushing worklogs into Jira.

Useful flags:
  --check          just test the Jira connection and stop
  --no-jira        skip Jira entirely and build the report from Outlook alone
  --no-outlook     skip the calendar and only show what Jira already holds
  --config PATH    use a different config file
  --out PATH       override the output workbook path
"""

from __future__ import annotations

import argparse
import configparser
import datetime as dt
import os
import sys
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import outlook_csv
from jira_worklogs import JiraClient, JiraError, Worklog, hours_by_day

HERE = os.path.dirname(os.path.abspath(__file__))

DAY_NAMES = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=12, bold=True)
FILL_SHORT = PatternFill("solid", fgColor="FCE4D6")   # under target
FILL_OK = PatternFill("solid", fgColor="E2EFDA")      # on target
FILL_OVER = PatternFill("solid", fgColor="FFF2CC")    # over target
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")   # you fill this in
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_ME = "<FILL IN>"


# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------
def load_config(path: str) -> configparser.ConfigParser:
    if not os.path.exists(path):
        sys.exit(
            f"Config file not found: {path}\n"
            f"Copy config.ini next to this script and fill it in."
        )
    cfg = configparser.ConfigParser(inline_comment_prefixes=("#", ";"))
    cfg.read(path, encoding="utf-8")
    return cfg


def cfg_get(cfg, section, key, default=""):
    try:
        return cfg.get(section, key).strip()
    except Exception:
        return default


def cfg_bool(cfg, section, key, default=True):
    raw = cfg_get(cfg, section, key, "")
    if not raw:
        return default
    return raw.strip().lower() in ("true", "yes", "1", "on")


def cfg_float(cfg, section, key, default=0.0):
    try:
        return float(cfg_get(cfg, section, key, str(default)))
    except ValueError:
        return default


def cfg_int(cfg, section, key, default=0):
    try:
        return int(float(cfg_get(cfg, section, key, str(default))))
    except ValueError:
        return default


def cfg_list(cfg, section, key) -> List[str]:
    raw = cfg_get(cfg, section, key, "")
    return [p.strip() for p in raw.split(",") if p.strip()]


# --------------------------------------------------------------------------
# Dates
# --------------------------------------------------------------------------
def default_window() -> Tuple[dt.date, dt.date]:
    """First day of last month through today - covers 'the past month or two'."""
    today = dt.date.today()
    first_this_month = today.replace(day=1)
    last_month_end = first_this_month - dt.timedelta(days=1)
    return last_month_end.replace(day=1), today


def daterange(start: dt.date, end: dt.date):
    day = start
    while day <= end:
        yield day
        day += dt.timedelta(days=1)


def round_hours(hours: float, to_minutes: int) -> float:
    if to_minutes <= 0:
        return round(hours, 2)
    step = to_minutes / 60.0
    return round(round(hours / step) * step, 4)


# --------------------------------------------------------------------------
# The actual gap calculation
# --------------------------------------------------------------------------
class DayRow:
    def __init__(self, day: dt.date, is_working: bool, target: float):
        self.day = day
        self.is_working = is_working
        self.target = target if is_working else 0.0
        self.logged = 0.0
        self.meetings = 0.0
        self.meeting_overlap = 0.0
        self.meeting_list: List[outlook_csv.Meeting] = []

    @property
    def gap(self) -> float:
        return max(0.0, round(self.target - self.logged, 4))

    @property
    def suggest_meetings(self) -> float:
        return round(min(self.gap, self.meetings), 4)

    @property
    def suggest_other(self) -> float:
        return round(self.gap - self.suggest_meetings, 4)

    @property
    def status(self) -> str:
        if not self.is_working:
            return "Non-working day"
        if self.logged == 0:
            return "NOTHING LOGGED"
        if self.gap > 0.01:
            return "SHORT"
        if self.logged - self.target > 0.01:
            return "OVER"
        return "OK"


def build_days(
    date_from: dt.date,
    date_to: dt.date,
    worklogs: List[Worklog],
    meetings: List[outlook_csv.Meeting],
    target: float,
    work_days: List[str],
    holidays: List[dt.date],
) -> List[DayRow]:
    logged = hours_by_day(worklogs)
    merged = outlook_csv.merged_hours_by_day(meetings)
    overlaps = outlook_csv.overlap_hours_by_day(meetings)

    meetings_by_day: Dict[dt.date, List[outlook_csv.Meeting]] = {}
    for m in meetings:
        meetings_by_day.setdefault(m.date, []).append(m)

    rows: List[DayRow] = []
    for day in daterange(date_from, date_to):
        is_working = (
            DAY_NAMES[day.weekday()] in work_days and day not in holidays
        )
        row = DayRow(day, is_working, target)
        row.logged = round(logged.get(day, 0.0), 4)
        row.meetings = round(merged.get(day, 0.0), 4)
        row.meeting_overlap = round(overlaps.get(day, 0.0), 4)
        row.meeting_list = sorted(meetings_by_day.get(day, []), key=lambda m: m.start)
        rows.append(row)
    return rows


def build_upload_rows(
    days: List[DayRow],
    meeting_issue: str,
    work_issue: str,
    round_to: int,
    include_non_working: bool = False,
) -> List[dict]:
    """Rows ready to be pushed into Jira - meetings first, then the remainder.

    Meetings are allocated in chronological order until the day's gap is used
    up, so a day already 6 hours logged with 4 hours of calls produces 2 hours
    of meeting entries, not 4. You never end up over 8.
    """
    out: List[dict] = []
    for row in days:
        if not row.is_working and not include_non_working:
            continue
        if row.gap <= 0.01:
            continue

        remaining = row.suggest_meetings
        for m in row.meeting_list:
            if remaining <= 0.01:
                break
            take = round_hours(min(m.hours, remaining), round_to)
            if take <= 0.01:
                continue
            remaining = round(remaining - take, 4)
            out.append(
                {
                    "issue_key": meeting_issue or FILL_ME,
                    "date": row.day,
                    "hours": take,
                    "comment": f"{m.subject} ({m.time_range})",
                    "source": "Outlook meeting",
                }
            )

        other = round_hours(row.suggest_other, round_to)
        if other > 0.01:
            out.append(
                {
                    "issue_key": work_issue or FILL_ME,
                    "date": row.day,
                    "hours": other,
                    "comment": FILL_ME if not work_issue else "Development / analysis work",
                    "source": "Remaining hours",
                }
            )
    return out


# --------------------------------------------------------------------------
# Excel output
# --------------------------------------------------------------------------
def _style_header(ws, headers: List[str], widths: List[int], row: int = 1):
    for col, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _write_row(ws, row_idx: int, values: list, fill: Optional[PatternFill] = None):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.font = BODY_FONT
        cell.border = BORDER
        if isinstance(value, float):
            cell.number_format = "0.00"
        if isinstance(value, dt.date):
            cell.number_format = "yyyy-mm-dd"
        if fill:
            cell.fill = fill


def write_workbook(
    path: str,
    days: List[DayRow],
    worklogs: List[Worklog],
    meetings: List[outlook_csv.Meeting],
    upload_rows: List[dict],
    meta: dict,
) -> None:
    wb = Workbook()

    # ---- Read Me ---------------------------------------------------------
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95
    ws["A1"] = "Jira timesheet gap report"
    ws["A1"].font = TITLE_FONT
    notes = [
        ("Generated", meta.get("generated", "")),
        ("Window", f"{meta.get('date_from')} to {meta.get('date_to')}"),
        ("Jira account", meta.get("jira_user", "(Jira not queried)")),
        ("Target hours per day", meta.get("target", "")),
        ("Working days", meta.get("work_days", "")),
        ("Outlook file", meta.get("outlook_path", "(not used)")),
        ("", ""),
        ("Daily Summary", "One row per day. 'Gap' is what is still missing after "
                          "what Jira already holds. Orange = short, green = on "
                          "target, yellow = more than target already logged."),
        ("Jira Worklogs", "Every entry already sitting in Jira for you. Read-only "
                          "evidence - do not edit."),
        ("Outlook Meetings", "Every meeting parsed from the calendar export."),
        ("Upload Template", "THE SHEET YOU ACT ON. Fill any yellow cells, delete "
                            "rows you do not want, then feed it to your existing "
                            "Jira upload script."),
        ("", ""),
        ("Meeting hours note", "Overlapping meetings are counted once. If two calls "
                               "sit on top of each other from 10-11, that is 1 hour "
                               "of your day, not 2."),
        ("Allocation rule", "Meetings are allocated first, oldest to newest, only up "
                            "to the size of the gap. Whatever is left over becomes a "
                            "single 'remaining hours' row for you to assign."),
        ("Nothing was written", "This report is read-only against Jira. No worklog "
                                "was created or changed by generating it."),
    ]
    for i, (label, text) in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
        c = ws.cell(row=i, column=2, value=text)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Daily Summary ---------------------------------------------------
    ws = wb.create_sheet("Daily Summary")
    headers = ["Date", "Day", "Working day", "Target hrs", "Logged in Jira",
               "Meeting hrs", "Gap", "Suggest: meetings", "Suggest: other", "Status"]
    _style_header(ws, headers, [12, 6, 12, 11, 14, 12, 9, 16, 14, 16])

    r = 2
    for row in days:
        fill = None
        if row.is_working:
            fill = {"SHORT": FILL_SHORT, "NOTHING LOGGED": FILL_SHORT,
                    "OVER": FILL_OVER, "OK": FILL_OK}.get(row.status)
        _write_row(ws, r, [
            row.day,
            DAY_NAMES[row.day.weekday()].title(),
            "Yes" if row.is_working else "No",
            float(row.target),
            float(row.logged),
            float(row.meetings),
            float(row.gap),
            float(row.suggest_meetings),
            float(row.suggest_other),
            row.status,
        ], fill)
        r += 1

    total_target = sum(d.target for d in days)
    total_logged = sum(d.logged for d in days)
    total_gap = sum(d.gap for d in days)
    _write_row(ws, r + 1, ["TOTAL", "", "", float(total_target), float(total_logged),
                           float(sum(d.meetings for d in days)), float(total_gap),
                           float(sum(d.suggest_meetings for d in days)),
                           float(sum(d.suggest_other for d in days)), ""])
    for col in range(1, 11):
        ws.cell(row=r + 1, column=col).font = Font(name="Arial", size=10, bold=True)

    # ---- Jira Worklogs ---------------------------------------------------
    ws = wb.create_sheet("Jira Worklogs")
    _style_header(ws, ["Date", "Issue key", "Summary", "Hours", "Comment", "Worklog ID"],
                  [12, 14, 45, 9, 55, 14])
    for i, wl in enumerate(worklogs, start=2):
        _write_row(ws, i, [wl.date, wl.issue_key, wl.issue_summary,
                           float(wl.hours), wl.comment, wl.worklog_id])

    # ---- Outlook Meetings ------------------------------------------------
    ws = wb.create_sheet("Outlook Meetings")
    _style_header(ws, ["Date", "Start", "End", "Hours", "Subject", "Organizer",
                       "Show as", "Location"],
                  [12, 8, 8, 8, 50, 26, 11, 26])
    for i, m in enumerate(meetings, start=2):
        _write_row(ws, i, [m.date, f"{m.start:%H:%M}", f"{m.end:%H:%M}",
                           float(m.hours), m.subject, m.organizer, m.show_as,
                           m.location])

    # ---- Upload Template -------------------------------------------------
    ws = wb.create_sheet("Upload Template")
    _style_header(ws, ["Issue Key", "Work Date", "Hours", "Comment", "Source"],
                  [16, 12, 9, 70, 18])
    ws.cell(row=1, column=6, value="<- rename these headers to match your existing "
                                   "upload script if they differ")
    ws.cell(row=1, column=6).font = Font(name="Arial", size=9, italic=True)

    for i, up in enumerate(upload_rows, start=2):
        _write_row(ws, i, [up["issue_key"], up["date"], float(up["hours"]),
                           up["comment"], up["source"]])
        for col in (1, 4):
            if ws.cell(row=i, column=col).value == FILL_ME:
                ws.cell(row=i, column=col).fill = FILL_INPUT

    if not upload_rows:
        _write_row(ws, 2, ["", None, None,
                           "Nothing to log - every working day already meets target.",
                           ""])

    wb.save(path)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def main() -> int:
    default_from, default_to = default_window()

    ap = argparse.ArgumentParser(description="Find the gaps in your Jira timesheet.")
    ap.add_argument("--config", default=os.path.join(HERE, "config.ini"))
    ap.add_argument("--from", dest="date_from", default=default_from.isoformat(),
                    help="YYYY-MM-DD (default: first day of last month)")
    ap.add_argument("--to", dest="date_to", default=default_to.isoformat(),
                    help="YYYY-MM-DD (default: today)")
    ap.add_argument("--outlook", default=None, help="Path to the Outlook CSV export")
    ap.add_argument("--out", default=None, help="Path for the Excel report")
    ap.add_argument("--check", action="store_true", help="Test the Jira connection only")
    ap.add_argument("--no-jira", action="store_true")
    ap.add_argument("--no-outlook", action="store_true")
    args = ap.parse_args()

    cfg = load_config(args.config)

    try:
        date_from = dt.date.fromisoformat(args.date_from)
        date_to = dt.date.fromisoformat(args.date_to)
    except ValueError:
        sys.exit("Dates must look like 2026-08-01")
    if date_to < date_from:
        sys.exit("--to is before --from")

    # ---- Jira ------------------------------------------------------------
    token = os.environ.get("JIRA_API_TOKEN") or cfg_get(cfg, "jira", "api_token")
    jira_user = None
    worklogs: List[Worklog] = []

    if args.check or not args.no_jira:
        try:
            client = JiraClient(
                base_url=cfg_get(cfg, "jira", "base_url"),
                deployment=cfg_get(cfg, "jira", "deployment", "cloud"),
                email=cfg_get(cfg, "jira", "email"),
                token=token,
                verify_ssl=cfg_bool(cfg, "jira", "verify_ssl", True),
            )
            jira_user = client.my_identity()
            print(f"Connected to Jira as: {jira_user}")
            if args.check:
                print("Connection is fine. Re-run without --check to build the report.")
                return 0
            print(f"Reading worklogs {date_from} -> {date_to} ...")
            worklogs = client.fetch_my_worklogs(date_from, date_to)
            print(f"  {len(worklogs)} worklog entries, "
                  f"{sum(w.hours for w in worklogs):.2f} hours total.")
        except JiraError as exc:
            if args.check:
                sys.exit(f"Jira check failed:\n  {exc}")
            print(f"\n!! Jira lookup failed: {exc}")
            print("!! Carrying on with Outlook only - the report will assume "
                  "nothing is logged yet.\n")

    # ---- Outlook ---------------------------------------------------------
    meetings: List[outlook_csv.Meeting] = []
    outlook_path = args.outlook or cfg_get(cfg, "outlook", "csv_path")
    if not args.no_outlook and outlook_path:
        if not os.path.exists(outlook_path):
            print(f"!! Outlook CSV not found at {outlook_path} - skipping meetings.")
            outlook_path = ""
        else:
            print(f"Reading calendar export: {outlook_path}")
            meetings, warns = outlook_csv.parse_outlook_csv(
                outlook_path,
                date_from=date_from,
                date_to=date_to,
                date_format=cfg_get(cfg, "outlook", "date_format", "auto"),
                skip_all_day=cfg_bool(cfg, "outlook", "skip_all_day", True),
                skip_free=cfg_bool(cfg, "outlook", "skip_free", True),
                skip_tentative=cfg_bool(cfg, "outlook", "skip_tentative", False),
                skip_cancelled=cfg_bool(cfg, "outlook", "skip_cancelled", True),
                min_minutes=cfg_int(cfg, "outlook", "min_meeting_minutes", 5),
                exclude_subjects=cfg_list(cfg, "outlook", "exclude_subjects"),
            )
            for w in warns:
                print(f"  note: {w}")
            merged_total = sum(outlook_csv.merged_hours_by_day(meetings).values())
            print(f"  {len(meetings)} meetings in window, "
                  f"{merged_total:.2f} hours after merging overlaps.")
    else:
        outlook_path = ""

    # ---- Crunch ----------------------------------------------------------
    target = cfg_float(cfg, "timesheet", "target_hours_per_day", 8.0)
    work_days = [d.upper()[:3] for d in cfg_list(cfg, "timesheet", "work_days")] or \
                ["MON", "TUE", "WED", "THU", "FRI"]
    holidays = []
    for h in cfg_list(cfg, "timesheet", "holidays"):
        try:
            holidays.append(dt.date.fromisoformat(h))
        except ValueError:
            print(f"  note: ignoring unreadable holiday '{h}'")

    days = build_days(date_from, date_to, worklogs, meetings, target, work_days, holidays)
    upload_rows = build_upload_rows(
        days,
        meeting_issue=cfg_get(cfg, "timesheet", "default_meeting_issue"),
        work_issue=cfg_get(cfg, "timesheet", "default_work_issue"),
        round_to=cfg_int(cfg, "timesheet", "round_to_minutes", 15),
    )

    out_path = args.out or cfg_get(cfg, "output", "excel_path") or \
        os.path.join(HERE, f"timesheet_gaps_{date_from}_{date_to}.xlsx")

    write_workbook(
        out_path, days, worklogs, meetings, upload_rows,
        meta={
            "generated": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "jira_user": jira_user or "(Jira not queried)",
            "target": target,
            "work_days": ", ".join(work_days),
            "outlook_path": outlook_path or "(not used)",
        },
    )

    short_days = [d for d in days if d.is_working and d.gap > 0.01]
    print("\n" + "=" * 62)
    print(f"Working days in window : {sum(1 for d in days if d.is_working)}")
    print(f"Days short of target   : {len(short_days)}")
    print(f"Hours missing in total : {sum(d.gap for d in days):.2f}")
    print(f"Rows ready to upload   : {len(upload_rows)}")
    print(f"Report written to      : {out_path}")
    print("=" * 62)
    return 0


if __name__ == "__main__":
    sys.exit(main())
